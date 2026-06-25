# This is to sum numbers within parenthesis, in this case denoting sum of the planned allocation
import os
import subprocess
import openpyxl as xl
from pathlib import Path

def to_normalize(text):
    return str(text).lower().replace(' ', '').replace(',', '')

def select_excel_file(title="Please Select Excel File"):
    # macOS原生文件选择弹窗
    script = f'''
    try
        set theFile to choose file with prompt "{title}" of type {{"xlsx", "xls", "xlsm"}}
        return POSIX path of theFile
    on error
        return ""
    end try
    '''
    result = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True, text=True, encoding="utf-8"
    )
    file_path = result.stdout.strip()
    if not file_path or not os.path.exists(file_path):
        return None
    return file_path

def show_macos_alert(title, message):
    script = f'''
    display dialog "{message}" with title "{title}" buttons {{"OK"}} default button 1
    '''
    subprocess.run(["osascript", "-e", script], capture_output=True, text=True, encoding="utf-8")

def to_dic_head_index(sheet, row = 1): #输出表头的index字典（从1开始）
    index_lib_0 = {}
    for index, cell in enumerate(sheet[row]):
        if cell.value is not None and cell.value != '':
            index_lib_0[to_normalize(cell.value)] = index + 1
    return index_lib_0

def parse_and_sum_allocation_string(s):
    """
    解析 "64(1),138(2),379(1)" 这种格式的字符串
    返回括号内数字和
    """
    result = 0
    if not s:  # 处理空值/None
        return result
    # 按逗号分割每一项
    items = s.strip().split(',')
    for item in items:
        item = item.strip()
        if not item:
            continue
        try:
            # 拆分 index 和 qty
            index_part, qty_part = item.split('(')
            qty = int(qty_part.rstrip(')'))
            result += qty
        except (ValueError, IndexError):
            #here是抓住ValueError, IndexError这两种错误吗？yes
            # 格式错误时跳过并打印提示
            err_msg = f"Invalid format data: {item}\nThis row will be skipped"
            show_macos_alert(title="Data Format Warning", message=err_msg)
            print(f"⚠️ 格式错误，跳过无效数据：{item}")
            continue
    return result

if __name__ == "__main__":
    the_file_path = select_excel_file(title="Please Select the Build Matrix Config Excel File")
    if not the_file_path:
        show_macos_alert("Operation Cancelled", "No Excel file selected, program exit")
        exit(0)
    file_p = Path(the_file_path)
    save_dir = file_p.parent
    wb = xl.load_workbook(filename=the_file_path)
    ws = wb[wb.sheetnames[0]]
    the_flat_header = to_dic_head_index(ws, row = 2)
    for row in range(3, ws.max_row + 1):
        target_cell = ws.cell(row=row, column=the_flat_header['plannedallocqty'])
        source_cell = ws.cell(row=row, column=the_flat_header['allocationindex'])
        if source_cell.value:
            the_sum = parse_and_sum_allocation_string(source_cell.value)
            target_cell.value = the_sum
    wb.save(filename=the_file_path)






